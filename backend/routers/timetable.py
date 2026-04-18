"""
Timetable generation + query endpoints.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional, Dict
import io

from database import get_db
from models import Timetable, TimetableSlot, Section, Subject, Teacher, Room
from schemas import TimetableOut, TimetableSlotOut, GenerateRequest, TimetableUpdate
from auth import require_role, get_current_user
from solver import generate_timetable

router = APIRouter(prefix="/api/timetable", tags=["Timetable"])


def _slot_to_out(s: TimetableSlot) -> TimetableSlotOut:
    return TimetableSlotOut(
        id=s.id, day=s.day, slot_index=s.slot_index,
        section_id=s.section_id,
        subject_id=s.subject_id, teacher_id=s.teacher_id,
        room_id=s.room_id, is_lab=s.is_lab,
        lab_engineer_id=s.lab_engineer_id,
        is_break=s.is_break, label=s.label,
        subject_code=s.subject.code if s.subject else None,
        subject_name=s.subject.full_name if s.subject else None,
        credit_hours=s.subject.theory_credits if s.subject else None,
        theory_credits=s.subject.theory_credits if s.subject else None,
        lab_credits=s.subject.lab_credits if s.subject else None,
        teacher_name=s.teacher.name if s.teacher else None,
        room_name=s.room.name if s.room else None,
        section_name=s.section.display_name if s.section else None,
        lab_engineer_name=s.lab_engineer.name if s.lab_engineer else None,
    )


@router.get("/list")
def list_timetables(db: Session = Depends(get_db), user=Depends(get_current_user)):
    q = db.query(Timetable)
    if user.role == "super_admin" or user.role == "vc":
        pass  # Can view all departments
    elif user.role == "program_admin":
        # Program admins see only their department's timetables
        if user.department_id:
            q = q.filter(Timetable.department_id == user.department_id)
    elif user.role == "teacher":
        # Teachers can see ALL timetables (for cross-department teaching)
        pass  # No filter - teachers can view all departments
    else:
        # Clerks: see their dept only
        if user.department_id:
            q = q.filter(Timetable.department_id == user.department_id)
            
    tts = q.order_by(Timetable.created_at.desc()).all()
    return [{
        "id": t.id,
        "name": t.name,
        "created_at": str(t.created_at),
        "status": t.status,
        "semester_info": t.semester_info,
        "created_by_id": t.created_by_id,
        "class_duration": t.class_duration,
        "break_start_time": t.break_start_time,
        "break_end_time": t.break_end_time,
        "max_slots_per_day": t.max_slots_per_day,
        "break_slot": t.break_slot,
        "friday_has_break": t.friday_has_break,
        "department_id": t.department_id,
        "session_id": t.session_id
    } for t in tts]


@router.post("/generate")
def generate(req: GenerateRequest, db: Session = Depends(get_db),
             user=Depends(require_role("super_admin", "program_admin"))):
    """Run the OR-Tools solver and store the result. Clerk cannot generate."""
    print(f"\n[GENERATE REQUEST] User: {user.username}, Session: {req.session_id}, Batches: {req.batch_ids}")
    print(f"[GENERATE REQUEST] Sequential Mode: {req.sequential_mode}")
    print(f"[GENERATE REQUEST] max_slots_per_day: {req.max_slots_per_day}")
    print(f"[GENERATE REQUEST] max_slots_friday: {req.max_slots_friday}")
    print(f"[GENERATE REQUEST] break_slot: {req.break_slot}")
    print(f"[GENERATE REQUEST] start_time: {req.start_time}")
    
    # Determine department_id: use user's dept if available,
    # otherwise derive from session's assignments (needed for super_admin)
    dept_id = user.department_id
    if not dept_id and req.session_id:
        from models import Assignment, Subject
        sample = db.query(Assignment).filter(
            Assignment.session_id == req.session_id
        ).first()
        if sample and sample.batch:
            dept_id = sample.batch.department_id
    
    print(f"[GENERATE REQUEST] Department ID: {dept_id}")
    
    try:
        # If incrementing an existing timetable, inherit settings from it
        target_timetable_id = req.timetable_id
        if target_timetable_id:
             tt_existing = db.query(Timetable).filter(Timetable.id == target_timetable_id).first()
             if tt_existing:
                 # Override request values with existing timetable settings for consistency
                 req.class_duration = tt_existing.class_duration
                 req.break_slot = tt_existing.break_slot
                 req.break_start_time = tt_existing.break_start_time
                 req.break_end_time = tt_existing.break_end_time
                 req.max_slots_per_day = tt_existing.max_slots_per_day
                 req.friday_has_break = tt_existing.friday_has_break

        # Check if sequential mode is requested
        if req.sequential_mode and req.batch_ids and len(req.batch_ids) > 1:
            # Use sequential batch-wise generation
            from solver import generate_timetable_sequential
            
            tt = generate_timetable_sequential(
                db=db,
                name=req.name,
                semester_info=req.semester_info,
                user_id=user.id,
                target_dept_id=dept_id,
                batch_ids_ordered=req.batch_ids,
                session_id=req.session_id,
                extra_classes_per_subject=req.extra_classes_per_subject,
                class_duration=req.class_duration,
                start_time=req.start_time,
                break_slot=req.break_slot,
                break_start_time=req.break_start_time,
                break_end_time=req.break_end_time,
                max_slots_per_day=req.max_slots_per_day,
                max_slots_friday=req.max_slots_friday,
                morning_lab_section_ids=req.morning_lab_section_ids,
                friday_has_break=req.friday_has_break,
                allow_friday_labs=req.allow_friday_labs,
                prefer_early_dismissal=req.prefer_early_dismissal,
                lab_is_last=req.lab_is_last,
                uniform_lab_start_batch_ids=req.uniform_lab_start_batch_ids
            )
        else:
            # Use regular generation (all batches at once)
            tt = generate_timetable(
                db, req.name, req.semester_info, user_id=user.id, 
                target_dept_id=dept_id,
                session_id=req.session_id,
                batch_ids=req.batch_ids,
                extra_classes_per_subject=req.extra_classes_per_subject,
                class_duration=req.class_duration,
                start_time=req.start_time,
                break_slot=req.break_slot,
                break_start_time=req.break_start_time,
                break_end_time=req.break_end_time,
                max_slots_per_day=req.max_slots_per_day,
                max_slots_friday=req.max_slots_friday,
                morning_lab_section_ids=req.morning_lab_section_ids,
                friday_has_break=req.friday_has_break,
                allow_friday_labs=req.allow_friday_labs,
                prefer_early_dismissal=req.prefer_early_dismissal,
                lab_is_last=req.lab_is_last,
                uniform_lab_start_batch_ids=req.uniform_lab_start_batch_ids,
                timetable_id=target_timetable_id
            )

    except ValueError as first_error:
        # ── Fallback: retry with morning labs for all sections ─────────
        # Only triggered when: (a) no sections were manually selected for morning,
        # AND (b) the first solve was INFEASIBLE.
        # This is a last-resort to avoid showing an error to the user.
        error_message = str(first_error)
        print(f"[GENERATION ERROR] {error_message}")
        
        # Don't retry fallback for sequential mode - just fail with detailed error
        if req.sequential_mode:
            raise HTTPException(status_code=400, detail=error_message)
        
        if not req.morning_lab_section_ids:
            print("[FALLBACK] First solve infeasible with no morning labs — retrying with morning labs as fallback")
            try:
                tt = generate_timetable(
                    db, req.name + " (Morning Fallback)", req.semester_info, user_id=user.id,
                    target_dept_id=dept_id,
                    session_id=req.session_id,
                    batch_ids=req.batch_ids,
                    extra_classes_per_subject=req.extra_classes_per_subject,
                    class_duration=req.class_duration,
                    break_slot=req.break_slot,
                    break_start_time=req.break_start_time,
                    break_end_time=req.break_end_time,
                    max_slots_per_day=req.max_slots_per_day,
                    max_slots_friday=req.max_slots_friday,
                    morning_lab_section_ids=None,  # None = all sections allowed morning labs
                    friday_has_break=req.friday_has_break,
                    allow_friday_labs=req.allow_friday_labs,
                    prefer_early_dismissal=req.prefer_early_dismissal,
                    lab_is_last=req.lab_is_last,
                    uniform_lab_start_batch_ids=req.uniform_lab_start_batch_ids,
                    timetable_id=target_timetable_id
                )
                print("[FALLBACK] Succeeded with morning labs fallback")
            except ValueError as fallback_error:
                # Both attempts failed — surface the FALLBACK error (more detailed)
                fallback_message = str(fallback_error)
                print(f"[FALLBACK ERROR] {fallback_message}")
                raise HTTPException(status_code=400, detail=fallback_message)
        else:
            # Sections were manually configured — don't fallback, show the error
            raise HTTPException(status_code=400, detail=error_message)
    except Exception as e:
        # Catch any other unexpected errors
        error_message = str(e)
        print(f"[UNEXPECTED ERROR] {error_message}")
        raise HTTPException(status_code=500, detail=f"Unexpected error during generation: {error_message}")

    # ── Success path ──────────────────────────────────────────────────
    # AUTO-ARCHIVE: Only archive previous timetables AFTER successful generation.
    # If generation fails, old timetables stay intact so the user can still view them.
    if dept_id:
        old_timetables = db.query(Timetable).filter(
            Timetable.department_id == dept_id,
            Timetable.status.in_(["generated", "active"]),
            Timetable.id != tt.id  # keep the newly created one
        ).all()
        archived_count = 0
        for old_tt in old_timetables:
            old_tt.status = "archived"
            archived_count += 1
        if archived_count > 0:
            db.commit()
            print(f"[AUTO-ARCHIVE] Archived {archived_count} old timetable(s) for department {dept_id}")

    slot_count = db.query(TimetableSlot).filter(TimetableSlot.timetable_id == tt.id).count()
    return {"id": tt.id, "name": tt.name, "status": tt.status,
            "slot_count": slot_count}


@router.delete("/{tt_id}")
def delete_timetable(tt_id: int, db: Session = Depends(get_db),
                     user=Depends(get_current_user)):
    """Delete a timetable. Requires super_admin or program_admin with department permission."""
    print(f"DEBUG: Delete attempt by {user.username} (Role: {user.role}, Dept: {user.department_id}) for TT: {tt_id}")
    if user.role not in ("super_admin", "program_admin"):
        print(f"DEBUG: Delete failed - Invalid role: {user.role}")
        raise HTTPException(403, "Insufficient permissions")
    
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        print(f"DEBUG: Delete failed - TT not found: {tt_id}")
        raise HTTPException(404, "Timetable not found")
    
    # Permission check: super_admin always allowed; 
    # program_admin needs flag. 
    # If tt.department_id is None, it's considered "Global/Shared" and allowed for any admin with the flag.
    if user.role == "program_admin":
        if not user.can_delete_timetable:
            print(f"DEBUG: Delete failed - missing can_delete_timetable flag for {user.username}")
            raise HTTPException(403, "You do not have permission to delete timetables")
        if tt.department_id and tt.department_id != user.department_id:
             print(f"DEBUG: Delete failed - department mismatch. TT: {tt.department_id}, User: {user.department_id}")
             raise HTTPException(403, "You can only delete timetables belonging to your department")
        
    db.delete(tt)
    db.commit()
    print(f"DEBUG: Delete successful for TT: {tt_id}")
    return {"ok": True}


@router.delete("/{tt_id}/batch/{batch_id}")
def delete_batch_from_timetable(tt_id: int, batch_id: int, db: Session = Depends(get_db),
                                user=Depends(require_role("super_admin", "program_admin"))):
    """Delete all slots for a specific batch from an existing timetable."""
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
        
    # Permission check
    if user.role == "program_admin":
        if tt.department_id and tt.department_id != user.department_id:
            raise HTTPException(403, "You can only manage timetables belonging to your department")

    from models import Section
    db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == tt_id,
        TimetableSlot.section_id.in_(
            db.query(Section.id).filter(Section.batch_id == batch_id)
        )
    ).delete(synchronize_session=False)
    
    db.commit()
    return {"ok": True}


@router.patch("/{tt_id}")
def update_timetable(tt_id: int, req: TimetableUpdate, db: Session = Depends(get_db),
                     user=Depends(get_current_user)):
    """Update timetable name or status. Requires same permissions as deletion."""
    print(f"DEBUG: Update attempt by {user.username} for TT: {tt_id}")
    if user.role not in ("super_admin", "program_admin"):
        raise HTTPException(403, "Insufficient permissions")
    
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
    
    # Permission check: super_admin always allowed; program_admin needs flag and dept match
    if user.role == "program_admin":
        if not user.can_delete_timetable: # Reusing this flag as a 'can manage' flag for now
            raise HTTPException(403, "You do not have permission to manage timetables")
        if tt.department_id and tt.department_id != user.department_id:
             raise HTTPException(403, "You can only manage timetables belonging to your department")
    
    if req.name is not None:
        tt.name = req.name
    if req.status is not None:
        tt.status = req.status
        
    db.commit()
    db.refresh(tt)
    return {"ok": True, "status": tt.status, "name": tt.name}


@router.post("/create")
def create_timetable(data: dict, db: Session = Depends(get_db),
                     user=Depends(require_role("super_admin", "program_admin"))):
    """Create a new empty timetable for manual editing."""
    name = data.get("name")
    department_id = data.get("department_id")
    session_id = data.get("session_id")
    
    if not name:
        raise HTTPException(400, "Timetable name is required")
    if not department_id:
        raise HTTPException(400, "Department is required")
    
    # Permission check for program_admin
    if user.role == "program_admin":
        if user.department_id != department_id:
            raise HTTPException(403, "You can only create timetables for your department")
    
    # Create new timetable with status="latest" so it's immediately visible
    tt = Timetable(
        name=name,
        department_id=department_id,
        session_id=session_id,
        created_by_id=user.id,
        status="latest",  # Changed from "draft" to "latest" for immediate visibility
        class_duration=60,
        break_start_time="11:00",
        break_end_time="11:30",
        max_slots_per_day=8,
        break_slot=2,
        friday_has_break=True
    )
    db.add(tt)
    db.commit()
    db.refresh(tt)
    
    return {"id": tt.id, "name": tt.name, "status": tt.status}


@router.post("/{tt_id}/mark-latest")
def mark_as_latest(tt_id: int, db: Session = Depends(get_db),
                   user=Depends(require_role("super_admin", "program_admin"))):
    """Mark a timetable as 'latest' (current active timetable)."""
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
    
    # Permission check
    if user.role == "program_admin":
        if tt.department_id and tt.department_id != user.department_id:
            raise HTTPException(403, "You can only manage timetables for your department")
    
    # Set all other timetables in this department to 'archived'
    db.query(Timetable).filter(
        Timetable.department_id == tt.department_id,
        Timetable.id != tt_id,
        Timetable.status == "latest"
    ).update({"status": "archived"})
    
    # Mark this one as latest
    tt.status = "latest"
    db.commit()
    
    return {"ok": True, "message": f"Timetable '{tt.name}' is now marked as Latest"}


@router.get("/{tt_id}")
def get_timetable(tt_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
    
    # Permission: program_admin, super_admin, and vc can view all; others only own dept
    if user.role not in ("super_admin", "program_admin", "vc") and user.department_id and tt.department_id != user.department_id:
        raise HTTPException(403, "Access denied to other department's timetable")
    
    slots = db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == tt_id
    ).order_by(TimetableSlot.section_id, TimetableSlot.day, TimetableSlot.slot_index).all()

    slot_list = []
    for s in slots:
        sec = db.query(Section).filter(Section.id == s.section_id).first()
        out = _slot_to_out(s)
        out.section_name = sec.display_name if sec else None
        slot_list.append(out)

    # Get all sections for this timetable
    section_ids = set(s.section_id for s in slots if s.section_id)
    sections = db.query(Section).options(joinedload(Section.batch)).filter(Section.id.in_(section_ids)).all() if section_ids else []
    sections_list = [{
        "id": sec.id,
        "name": sec.display_name,
        "display_name": sec.display_name,
        "batch_id": sec.batch_id,
        "batch_year": sec.batch.year if sec.batch else None
    } for sec in sections]

    return {
        "id": tt.id, "name": tt.name, "status": tt.status,
        "created_at": str(tt.created_at),
        "semester_info": tt.semester_info,
        "department_id": tt.department_id,
        "class_duration": tt.class_duration,
        "start_time": tt.start_time,
        "break_start_time": tt.break_start_time,
        "break_end_time": tt.break_end_time,
        "break_slot": tt.break_slot,
        "max_slots_per_day": tt.max_slots_per_day,
        "slots": [sl.model_dump() for sl in slot_list],
        "sections": sections_list,
        "friday_has_break": tt.friday_has_break,
        "max_slots_friday": tt.max_slots_friday
    }


@router.get("/{tt_id}/section/{section_id}")
def get_section_timetable(tt_id: int, section_id: int,
                          db: Session = Depends(get_db)):
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
    
    slots = db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == tt_id,
        TimetableSlot.section_id == section_id,
    ).order_by(TimetableSlot.day, TimetableSlot.slot_index).all()
    sec = db.query(Section).filter(Section.id == section_id).first()
    return {
        "section": sec.display_name if sec else str(section_id),
        "slots": [_slot_to_out(s).model_dump() for s in slots],
        "class_duration": tt.class_duration,
        "start_time": tt.start_time,
        "break_start_time": tt.break_start_time,
        "break_end_time": tt.break_end_time,
        "break_slot": tt.break_slot,
        "max_slots_per_day": tt.max_slots_per_day,
        "max_slots_friday": tt.max_slots_friday,
        "friday_has_break": tt.friday_has_break
    }


@router.get("/{tt_id}/teacher/{teacher_id}")
def get_teacher_timetable(tt_id: int, teacher_id: int,
                          db: Session = Depends(get_db)):
    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")
    
    slots = db.query(TimetableSlot).options(
        joinedload(TimetableSlot.subject),
        joinedload(TimetableSlot.teacher),
        joinedload(TimetableSlot.room),
        joinedload(TimetableSlot.section),
        joinedload(TimetableSlot.lab_engineer)
    ).filter(
        TimetableSlot.timetable_id == tt_id,
        (TimetableSlot.teacher_id == teacher_id) |
        (TimetableSlot.lab_engineer_id == teacher_id),
    ).order_by(TimetableSlot.day, TimetableSlot.slot_index).all()
    teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
    return {
        "teacher": teacher.name if teacher else str(teacher_id),
        "total_hours": len(slots),
        "slots": [_slot_to_out(s).model_dump() for s in slots],
        "class_duration": tt.class_duration,
        "start_time": tt.start_time,
        "break_start_time": tt.break_start_time,
        "break_end_time": tt.break_end_time,
        "break_slot": tt.break_slot,
        "max_slots_per_day": tt.max_slots_per_day,
        "max_slots_friday": tt.max_slots_friday,
        "friday_has_break": tt.friday_has_break
    }


@router.get("/{tt_id}/my-schedule")
def my_schedule(tt_id: int, db: Session = Depends(get_db),
                user=Depends(get_current_user)):
    if not user.teacher_id:
        raise HTTPException(400, "No teacher profile linked to this user")
    return get_teacher_timetable(tt_id, user.teacher_id, db)


@router.get("/{tt_id}/export/excel")
def export_excel(tt_id: int, db: Session = Depends(get_db)):
    """Export timetable as Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    tt = db.query(Timetable).filter(Timetable.id == tt_id).first()
    if not tt:
        raise HTTPException(404, "Timetable not found")

    slots = db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == tt_id
    ).all()

    # Group slots by section
    from collections import defaultdict
    section_slots = defaultdict(list)
    for s in slots:
        section_slots[s.section_id].append(s)

    DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    SLOT_TIMES = ["8:30-9:30", "9:30-10:30", "Break", "11:00-12:00",
                  "12:00-1:00", "1:00-2:00", "2:00-3:00", "3:00-4:00"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    row = 1

    for sec_id, sec_slots in sorted(section_slots.items()):
        sec = db.query(Section).filter(Section.id == sec_id).first()
        sec_name = sec.display_name if sec else str(sec_id)
        room = sec.room if sec else None
        room_name = room.name if room else ""

        ws.cell(row=row, column=1, value=f"{sec_name} - {room_name}")
        row += 1

        # Header
        ws.cell(row=row, column=1, value="Day / Time")
        for ci, t in enumerate(SLOT_TIMES):
            ws.cell(row=row, column=ci + 2, value=t)
        row += 1

        # Build grid
        grid = {}
        for s in sec_slots:
            grid[(s.day, s.slot_index)] = s

        for d in range(5):
            ws.cell(row=row, column=1, value=DAY_NAMES[d])
            for si in range(8):
                s = grid.get((d, si))
                if s:
                    if s.is_break:
                        label = "Break"
                    elif s.label:
                        label = s.label
                    elif s.subject:
                        label = s.subject.code
                        if s.is_lab:
                            label += " (Pr)"
                    else:
                        label = ""
                    ws.cell(row=row, column=si + 2, value=label)
            row += 1
        row += 1  # blank between sections

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=timetable_{tt_id}.xlsx"},
    )


# ── Manual Timetable Editing Endpoints ──────────────────────────

@router.post("/check-clash")
def check_clash(data: dict, db: Session = Depends(get_db)):
    """Check if adding a slot would create a teacher or room clash."""
    timetable_id = data.get('timetable_id')
    day = data.get('day')
    slot_index = data.get('slot_index')
    teacher_id = data.get('teacher_id')
    room_id = data.get('room_id')
    is_lab = data.get('is_lab', False)
    lab_engineer_id = data.get('lab_engineer_id')
    
    # Get all latest (non-archived) timetables
    latest_timetables = db.query(Timetable).filter(
        Timetable.status.in_(['generated', 'active'])
    ).all()
    
    clashes = []
    
    # For THEORY: Check teacher clash
    # For LAB: Check lab engineer clash (NOT teacher)
    if is_lab and lab_engineer_id:
        # Check lab engineer clash for labs
        for tt in latest_timetables:
            slots_to_check = [slot_index, slot_index + 1, slot_index + 2]
            
            for check_slot in slots_to_check:
                existing = db.query(TimetableSlot).filter(
                    TimetableSlot.timetable_id == tt.id,
                    TimetableSlot.day == day,
                    TimetableSlot.slot_index == check_slot,
                    TimetableSlot.lab_engineer_id == lab_engineer_id,
                    TimetableSlot.is_break == False
                ).first()
                
                if existing:
                    lab_eng = db.query(Teacher).filter(Teacher.id == lab_engineer_id).first()
                    section = db.query(Section).filter(Section.id == existing.section_id).first()
                    clashes.append(
                        f"Lab Engineer {lab_eng.name if lab_eng else lab_engineer_id} already supervising "
                        f"{section.display_name if section else 'another section'} at this time "
                        f"in timetable '{tt.name}'"
                    )
    elif teacher_id and not is_lab:
        # Check teacher clash for theory classes only
        for tt in latest_timetables:
            existing = db.query(TimetableSlot).filter(
                TimetableSlot.timetable_id == tt.id,
                TimetableSlot.day == day,
                TimetableSlot.slot_index == slot_index,
                TimetableSlot.teacher_id == teacher_id,
                TimetableSlot.is_break == False
            ).first()
            
            if existing:
                teacher = db.query(Teacher).filter(Teacher.id == teacher_id).first()
                section = db.query(Section).filter(Section.id == existing.section_id).first()
                clashes.append(
                    f"Teacher {teacher.name if teacher else teacher_id} already teaching "
                    f"{section.display_name if section else 'another section'} at this time "
                    f"in timetable '{tt.name}'"
                )
    
    # Check room clash in the same timetable
    if room_id:
        slots_to_check = [slot_index]
        if is_lab:
            slots_to_check = [slot_index, slot_index + 1, slot_index + 2]
        
        for check_slot in slots_to_check:
            existing = db.query(TimetableSlot).filter(
                TimetableSlot.timetable_id == timetable_id,
                TimetableSlot.day == day,
                TimetableSlot.slot_index == check_slot,
                TimetableSlot.room_id == room_id,
                TimetableSlot.is_break == False
            ).first()
            
            if existing:
                room = db.query(Room).filter(Room.id == room_id).first()
                section = db.query(Section).filter(Section.id == existing.section_id).first()
                clashes.append(
                    f"Room {room.name if room else room_id} already occupied by "
                    f"{section.display_name if section else 'another section'} at this time"
                )
    
    if clashes:
        return {
            "has_clash": True,
            "message": " | ".join(clashes)
        }
    
    return {"has_clash": False}


@router.post("/add-slot")
def add_slot(data: dict, db: Session = Depends(get_db),
             user=Depends(require_role("super_admin", "program_admin"))):
    """Add a single slot to a timetable."""
    timetable_id = data.get('timetable_id')
    section_id = data.get('section_id')
    day = data.get('day')
    slot_index = data.get('slot_index')
    subject_id = data.get('subject_id')
    teacher_id = data.get('teacher_id')
    room_id = data.get('room_id')
    is_lab = data.get('is_lab', False)
    lab_engineer_id = data.get('lab_engineer_id')
    
    # Check if slot already exists
    existing = db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == timetable_id,
        TimetableSlot.section_id == section_id,
        TimetableSlot.day == day,
        TimetableSlot.slot_index == slot_index
    ).first()
    
    if existing:
        raise HTTPException(400, "Slot already occupied")
    
    # Create slot(s)
    if is_lab:
        # Lab occupies 3 consecutive slots
        for offset in range(3):
            slot = TimetableSlot(
                timetable_id=timetable_id,
                section_id=section_id,
                day=day,
                slot_index=slot_index + offset,
                subject_id=subject_id,
                teacher_id=teacher_id,
                room_id=room_id,
                is_lab=True,
                lab_engineer_id=lab_engineer_id
            )
            db.add(slot)
    else:
        slot = TimetableSlot(
            timetable_id=timetable_id,
            section_id=section_id,
            day=day,
            slot_index=slot_index,
            subject_id=subject_id,
            teacher_id=teacher_id,
            room_id=room_id,
            is_lab=False
        )
        db.add(slot)
    
    db.commit()
    return {"success": True}


@router.post("/add-break")
def add_break(data: dict, db: Session = Depends(get_db),
              user=Depends(require_role("super_admin", "program_admin"))):
    """Add a break slot to a timetable."""
    timetable_id = data.get('timetable_id')
    section_id = data.get('section_id')
    day = data.get('day')
    slot_index = data.get('slot_index')
    
    # Check if slot already exists
    existing = db.query(TimetableSlot).filter(
        TimetableSlot.timetable_id == timetable_id,
        TimetableSlot.section_id == section_id,
        TimetableSlot.day == day,
        TimetableSlot.slot_index == slot_index
    ).first()
    
    if existing:
        return {"success": True}  # Already exists, no need to add
    
    # Create break slot
    slot = TimetableSlot(
        timetable_id=timetable_id,
        section_id=section_id,
        day=day,
        slot_index=slot_index,
        subject_id=None,
        teacher_id=None,
        room_id=None,
        is_lab=False,
        is_break=True
    )
    db.add(slot)
    db.commit()
    return {"success": True}


@router.delete("/slot/{slot_id}")
def delete_slot(slot_id: int, db: Session = Depends(get_db),
                user=Depends(require_role("super_admin", "program_admin"))):
    """Delete a single slot."""
    slot = db.query(TimetableSlot).filter(TimetableSlot.id == slot_id).first()
    if not slot:
        raise HTTPException(404, "Slot not found")
    
    # If it's a lab slot, delete all 3 consecutive slots
    if slot.is_lab:
        db.query(TimetableSlot).filter(
            TimetableSlot.timetable_id == slot.timetable_id,
            TimetableSlot.section_id == slot.section_id,
            TimetableSlot.day == slot.day,
            TimetableSlot.slot_index.in_([slot.slot_index, slot.slot_index + 1, slot.slot_index + 2]),
            TimetableSlot.subject_id == slot.subject_id
        ).delete()
    else:
        db.delete(slot)
    
    db.commit()
    return {"success": True}
